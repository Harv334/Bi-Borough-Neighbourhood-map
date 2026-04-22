#\!/usr/bin/env python3
"""
Replace `census_population` in lsoa_data.json and ward_data.json with ONS
mid-2024 population estimates (single-year-of-age LSOA dataset,
mid-2022 to mid-2024 release).

Usage:
    python scripts/patch_population_mid2024.py

The input xlsx must be at:
    raw_data/sapelsoasyoa20222024.xlsx

Source:  ONS "Population estimates for Lower layer Super Output Areas in
         England and Wales by single year of age and sex" (mid-2022 to
         mid-2024), sheet "Mid-2024 LSOA 2021".
"""
from __future__ import annotations

import json
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required - pip install openpyxl")

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "raw_data" / "sapelsoasyoa20222024.xlsx"
LSOA_JSON = REPO / "lsoa_data.json"
WARD_JSON = REPO / "ward_data.json"
INDEX_HTML = REPO / "index.html"

SHEET = "Mid-2024 LSOA 2021"
SOURCE_LABEL = "ONS mid-2024"
SOURCE_YEAR = "2024"


def _safe_write(path: Path, text: str) -> None:
    """Stage via /tmp then copy across virtiofs, verifying size each step.

    Direct writes to the mount have been observed to truncate large files
    silently. This helper rejects silent truncation with an assertion.
    """
    expected = len(text.encode("utf-8"))
    tmp = Path(tempfile.gettempdir()) / (path.name + ".staging")
    tmp.write_text(text, encoding="utf-8")
    assert tmp.stat().st_size == expected, (
        f"staging write short: {tmp.stat().st_size} vs {expected}"
    )
    # copy in 4MB chunks so we can detect short-write across the mount
    with open(tmp, "rb") as src, open(path, "wb") as dst:
        while chunk := src.read(4 * 1024 * 1024):
            dst.write(chunk)
    got = path.stat().st_size
    assert got == expected, (
        f"virtiofs truncated {path.name}: {got} vs {expected}"
    )
    tmp.unlink()


def load_pop_estimates() -> dict:
    """Parse the xlsx and return {lsoa_code: total_population}."""
    if not XLSX.exists():
        sys.exit(f"missing input: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"sheet not found: {SHEET}")
    ws = wb[SHEET]

    pop = {}
    header_seen = False
    code_col = total_col = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if not header_seen:
            if any(c == "LSOA 2021 Code" for c in row if isinstance(c, str)):
                cols = list(row)
                code_col = cols.index("LSOA 2021 Code")
                total_col = cols.index("Total")
                header_seen = True
            continue
        code = row[code_col]
        total = row[total_col]
        if isinstance(code, str) and code.startswith("E01") and isinstance(total, (int, float)):
            pop[code] = int(total)
    return pop


def patch_lsoa_json(pop):
    data = json.loads(LSOA_JSON.read_text(encoding="utf-8"))
    updated = missing = 0
    for code, entry in data.items():
        if code in pop:
            entry["census_population"] = pop[code]
            updated += 1
        else:
            missing += 1
    _safe_write(LSOA_JSON, json.dumps(data, indent=2))
    return updated, missing


def patch_ward_json(pop):
    """Re-aggregate ward census_population by summing LSOAs in each ward."""
    html = INDEX_HTML.read_text(encoding="utf-8")
    m = re.search(r"const LSOA_IMD = (\{.*?\});", html, re.DOTALL)
    if not m:
        sys.exit("LSOA_IMD constant not found in index.html")
    geo = json.loads(m.group(1))

    lsoa_by_ward = defaultdict(list)
    for feat in geo["features"]:
        p = feat["properties"]
        wc = p.get("ward_code")
        lc = p.get("code")
        if wc and lc:
            lsoa_by_ward[wc].append(lc)

    ward_doc = json.loads(WARD_JSON.read_text(encoding="utf-8"))
    wards = ward_doc["wards"]
    updated = skipped = 0
    for wc, w in wards.items():
        codes = lsoa_by_ward.get(wc, [])
        total = sum(pop[c] for c in codes if c in pop)
        if total > 0:
            w.setdefault("indicators", {})["census_population"] = total
            updated += 1
        else:
            skipped += 1

    meta = ward_doc.setdefault("metadata", {})
    meta["census_population_source"] = SOURCE_LABEL
    meta["census_population_year"] = SOURCE_YEAR

    _safe_write(WARD_JSON, json.dumps(ward_doc, indent=2))
    return updated, skipped


def patch_index_html_labels():
    html = INDEX_HTML.read_text(encoding="utf-8")
    pattern = re.compile(
        r'(census_population:\s*\{\s*src:\s*")[^"]+(",\s*yr:\s*")[^"]+(",)',
    )
    new_html, n = pattern.subn(
        r'\1' + SOURCE_LABEL + r'\2' + SOURCE_YEAR + r'\3',
        html,
        count=1,
    )
    if n:
        _safe_write(INDEX_HTML, new_html)
    return n


def main():
    print(f"reading {XLSX.name} ...", flush=True)
    pop = load_pop_estimates()
    print(f"  {len(pop):,} LSOA totals loaded from '{SHEET}'")

    print(f"patching {LSOA_JSON.name} ...", flush=True)
    u, m = patch_lsoa_json(pop)
    print(f"  updated {u:,} LSOAs, {m:,} not in ONS file")

    print(f"re-aggregating {WARD_JSON.name} ...", flush=True)
    u, s = patch_ward_json(pop)
    print(f"  updated {u} wards, {s} skipped (no LSOA match)")

    print(f"updating source label in {INDEX_HTML.name} ...", flush=True)
    n = patch_index_html_labels()
    print(f"  replaced {n} label entry")

    print("\ndone. census_population now reflects ONS mid-2024 estimates.")


if __name__ == "__main__":
    main()
